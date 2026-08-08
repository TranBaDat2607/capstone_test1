#!/usr/bin/env python3
"""
So sánh Graph-RAG với RAG thường trên cùng một tập câu bằng chứng.

Mỗi câu `evidence_text` trong sheet_A đi vào HAI nhánh, rồi so kết quả:

    evidence_text ─┬─→ [A] Graph-RAG : truy xuất evidence→claim trên resolved_graph.json
                   │                   (trục chỉ tiêu 2-hop + trùng token + cửa sổ thời gian)
                   └─→ [B] RAG thường: query → BM25 + dense + RRF + LLM rerank
                                       (ragtest/, corpus 8.956 câu claim)
                              ↓                     ↓
                          claim_A               claim_B
                              └────── so sánh ──────┘

**Nhãn quan hệ do CÙNG MỘT `Adjudicator` của step07 chấm cho cả hai nhánh** — cùng
prompt `ADJUDICATE_SYSTEM`, cùng model, temperature=0. Đây là điều kiện bắt buộc: nếu
mỗi nhánh dùng bộ chấm riêng thì chênh lệch nhãn lẫn lộn giữa "khác cơ chế truy xuất"
và "khác prompt", và số đo mất nghĩa (`AGENT_AB_EVALUATION.md` §3.1: prompt dùng chung
phải giữ nguyên văn).

Đo cái gì:
  1. **Word matching** — token P/R/F1, Jaccard, ROUGE-L giữa claim_A và claim_B
  2. **Ngữ nghĩa** — cosine trên vector nhúng của hai câu claim
  3. **Nhãn** — ma trận nhầm lẫn 3×3, tỷ lệ đồng thuận, Cohen's kappa
  4. **Kiểm định ghép cặp** — McNemar trên "nhánh nào hay kết luận không-irrelevant hơn"
  5. **Đối chứng âm** (mặc định bật) — ghép mỗi câu bằng chứng với một claim NGẪU NHIÊN
     rồi cho cùng Adjudicator chấm. Không có nó thì mọi con số ở trên vô nghĩa.

Vì sao đối chứng âm là bắt buộc chứ không phải tuỳ chọn: `AGENT_AB_EVALUATION.md` §2
gọi đây là "luật hai trục". Mọi metric đo *nhiều hơn* (coverage, tỷ lệ ra claim) phải
đi kèm một metric đo *có bịa không*, đo trên cùng lần chạy. Một nhánh trả nhiều claim
hơn có thể vì nó tìm ra bằng chứng thật, mà cũng có thể vì nó dễ dãi hơn — hai giả
thuyết đó cho ra cùng một con số coverage. Specificity trên cặp ngẫu nhiên là thứ tách
được chúng.

Giới hạn phải nói rõ khi đọc kết quả: **KHÔNG có ground truth.** sheet_A/sheet_B đều
220/220 chưa ai chấm. Nên đây đo hai hệ KHÁC NHAU ra sao, không đo hệ nào ĐÚNG hơn.
Repo cấm hẳn nhóm metric cần nhãn (`EVALUATION_WITHOUT_LABELS.md` §8: link precision,
kappa người–người, kappa người–LLM, recall vét cạn, accuracy/precision/recall đối với
sự thật) — không có cái nào trong số đó xuất hiện ở đây.

    python test3/compare_graphrag_vs_rag.py --limit 10     # chạy thử
    python test3/compare_graphrag_vs_rag.py                # đủ 220 dòng
    python test3/compare_graphrag_vs_rag.py --no-negative-control   # bỏ trục kỷ luật (không khuyến khích)
"""

from __future__ import annotations

import argparse
import contextlib
import csv
import io
import json
import random
import sys
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
for _p in (str(REPO_ROOT), str(REPO_ROOT / "src")):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from esg_kg.core.console import ensure_utf8_stdout  # noqa: E402

# Dùng lại phần dựng câu query đã làm ở test2 (evidence + description + valid_from) và
# phần tra ngược evidence_text → node. Import chứ không chép: chép sang đây là hai bản
# có thể lệch nhau về sau mà không ai biết.
from test2.build_evidence_claim_sheet import (  # noqa: E402
    build_query,
    build_text_index,
    row_fields,
)
from test3.compare_metrics import (  # noqa: E402
    bootstrap_ci,
    cosine,
    label_agreement,
    mcnemar_exact,
    text_similarity,
    wilson_ci,
)
from test3.graph_rag_arm import GraphArm  # noqa: E402

SHEET_A = REPO_ROOT / "evalu" / "out" / "annotation" / "sheet_A.json"
RESOLVED_GRAPH = REPO_ROOT / "graph_output" / "resolved" / "resolved_graph.json"

SHEET_COLUMNS = (
    "pair_id", "ticker", "evidence_text",
    "graph_claim", "graph_tier", "graph_relation",
    "rag_claim", "rag_relation",
    "token_f1", "token_precision", "token_recall", "jaccard", "rouge_l", "cosine",
    "labels_agree",
    "graph_claim_norerank", "graph_rerank_changed",
)


# --------------------------------------------------------------------------
def match_node_index(evidence_text: str, kind_label: Optional[str],
                     text_index: Dict[str, List[int]],
                     nodes: Sequence[Dict[str, Any]]) -> Optional[int]:
    """
    CHỈ SỐ của node ứng với một câu bằng chứng, hoặc None.

    Cùng logic `build_evidence_claim_sheet.match_node` nhưng trả chỉ số thay vì node.
    Phải là chỉ số, vì `GraphArm` đi theo cạnh mà cạnh trong resolved_graph.json trỏ
    node bằng chỉ số mảng. Tuyệt đối KHÔNG dùng `nodes.index(node)` để suy ra: nó so
    sánh dict theo giá trị, nên hai node trùng nội dung sẽ cho cùng một chỉ số — đúng
    cái bẫy `EVALUATION_WITHOUT_LABELS.md` §9.1 mô tả, sai ~50% mà không báo lỗi.
    """
    if not evidence_text:
        return None
    candidates = text_index.get(evidence_text)
    if not candidates:
        return None
    if kind_label is not None:
        candidates = [i for i in candidates if nodes[i].get("class") == kind_label]
    return candidates[0] if candidates else None


def adjudicate_label(adjud, claim_text: str, evidence_text: str, meta: str) -> str:
    """
    Một nhãn quan hệ cho cặp (claim, evidence), do Adjudicator của step07 chấm.

    Không có claim thì trả "" — KHÔNG trả "irrelevant". Hai chuyện đó khác nhau: một
    bên là hệ không truy xuất được gì, bên kia là hệ có truy xuất được nhưng phán là
    không liên quan. Gộp chúng lại sẽ khiến một nhánh phủ kém trông như một nhánh
    nghiêm khắc.
    """
    if not claim_text:
        return ""
    verdict = adjud.adjudicate(claim_text, evidence_text, meta)
    if not verdict:
        return ""
    value = verdict.get("verdict")
    return value if value in ("supports", "contradicts", "irrelevant") else ""


def write_outputs(rows: Sequence[Dict[str, Any]], summary: Dict[str, Any],
                  out_dir: Path, name: str) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)

    (out_dir / f"{name}.json").write_text(
        json.dumps({"summary": summary, "rows": list(rows)}, ensure_ascii=False, indent=2),
        encoding="utf-8")
    print(f"  {out_dir / (name + '.json')}")

    with open(out_dir / f"{name}.csv", "w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(SHEET_COLUMNS))
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in SHEET_COLUMNS})
    print(f"  {out_dir / (name + '.csv')}")

    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  (bỏ qua .xlsx — chưa cài openpyxl)")
        return

    book = openpyxl.Workbook()
    sheet = book.active
    sheet.title = "graph_vs_rag"
    sheet.append(list(SHEET_COLUMNS))
    for column in range(1, len(SHEET_COLUMNS) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.freeze_panes = "A2"
    for row in rows:
        sheet.append([row.get(k, "") for k in SHEET_COLUMNS])
    widths = {"evidence_text": 55, "graph_claim": 55, "rag_claim": 55, "pair_id": 14}
    for i, key in enumerate(SHEET_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(i)].width = widths.get(key, 14)
    for cells in sheet.iter_rows(min_row=2):
        for cell in cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    book.save(out_dir / f"{name}.xlsx")
    print(f"  {out_dir / (name + '.xlsx')}")


def render_report(summary: Dict[str, Any]) -> str:
    """Báo cáo Markdown — mọi con số đều kèm mẫu số và giới hạn của nó."""
    L: List[str] = []
    add = L.append
    add("# Graph-RAG vs RAG thường\n")
    add(f"- Số câu bằng chứng: **{summary['n_items']}**")
    add(f"- Model chấm nhãn (dùng chung cho cả hai nhánh): `{summary['model']}`")
    add("- **Không có ground truth** — sheet_A/sheet_B đều chưa ai chấm. Bảng dưới đo "
        "hai hệ KHÁC NHAU ra sao, không đo hệ nào ĐÚNG hơn.\n")

    add("## 1. Độ phủ (trục X — năng suất)\n")
    add("| nhánh | ra được claim | tỷ lệ |")
    add("|---|---|---|")
    for arm in ("graph", "rag"):
        k, n = summary[f"{arm}_answered"], summary["n_items"]
        add(f"| {arm} | {k}/{n} | {k / n:.1%} |" if n else f"| {arm} | 0/0 | – |")
    add("\n> Coverage một mình KHÔNG nói lên chất lượng: nhánh trả nhiều hơn có thể vì "
        "tìm được bằng chứng thật, cũng có thể chỉ vì dễ dãi hơn. Đọc kèm mục 4.\n")

    add("### 1a. Hai nhánh đã được cân bằng thế nào\n")
    add("| | nhánh A (graph) | nhánh B (rag) |")
    add("|---|---|---|")
    add("| sinh ứng viên | đồ thị: trục chỉ tiêu + trùng token | BM25 + dense, hợp nhất RRF |")
    add(f"| LLM rerank | {'✅ CÙNG reranker' if summary.get('reranked') else '❌ không'} "
        f"| {'✅ có' if summary.get('reranked') else '❌ không'} |")
    add("| chấm nhãn | ✅ cùng Adjudicator | ✅ cùng Adjudicator |")
    add("\n> Cân bằng reranker là bắt buộc. Nếu một nhánh có reranker còn nhánh kia không, "
        "chênh lệch đo được là chênh lệch **của reranker**, không phải của cơ chế truy xuất — "
        "đúng cái bẫy mà bài đánh giá hệ thống RAG vs GraphRAG (arXiv 2502.11371) xử lý bằng "
        "cách chuẩn hoá cấu hình và cân bằng ngân sách giữa hai hệ.\n")
    changed = summary.get("graph_rerank_changed")
    if changed is not None:
        add(f"- Reranker đổi claim top-1 của nhánh graph ở **{changed}** dòng "
            "(0 nghĩa là đồ thị hầu như chỉ trả về 1 ứng viên, reranker không có gì để sắp).\n")

    tiers = summary.get("graph_tiers") or {}
    ind, tok = tiers.get("indicator", 0), tiers.get("token_overlap", 0)
    add("### 1b. Nhánh graph dùng tầng nào\n")
    add(f"- tầng **trục chỉ tiêu** (nối cấu trúc 2 bước): **{ind}**")
    add(f"- tầng **trùng token** chủ đề tiếng Việt: **{tok}**\n")
    if ind == 0 and tok > 0:
        add("> ⚠️ **Tầng nối cấu trúc đóng góp 0.** Trên tập bằng chứng này, thứ làm nên "
            "chữ \"Graph\" trong Graph-RAG không kích hoạt lần nào, nên nhánh A thực chất "
            "đang là *trùng token tiếng Việt + khoanh theo doanh nghiệp + cửa sổ thời gian*. "
            "Nguyên nhân đã truy được: MediaReport theo schema không mang cạnh "
            "`measuredUnder`; KPIObservation phía tin tức không có cạnh đó (1.309 cạnh "
            "`measuredUnder` nằm ở KPI phía báo cáo); còn Penalty thì cả 85 node trỏ vào "
            "đúng một chỉ tiêu — *\"Tổng tiền phạt vi phạm môi trường\"* — mà không doanh "
            "nghiệp nào tuyên bố mình bị phạt, nên đầu claim của chỉ tiêu đó trống. Bản "
            "thân trục chỉ tiêu lành (292/464 claim có `alignsWithIndicator`, 95 chỉ tiêu "
            "có claim); chỉ là hai nửa không gặp nhau ở đúng tập bằng chứng này. "
            "**Đọc mọi số dưới đây với nhãn đúng đó.**\n")

    add("## 2. Hai câu claim giống nhau đến đâu\n")
    add("| metric | trung bình | KTC 95% (bootstrap) |")
    add("|---|---|---|")
    for key, label in (("token_f1", "token F1"), ("token_precision", "token precision"),
                       ("token_recall", "token recall"), ("jaccard", "Jaccard"),
                       ("rouge_l", "ROUGE-L"), ("cosine", "cosine (ngữ nghĩa)")):
        stat = summary["similarity"].get(key) or {}
        mean, ci = stat.get("mean"), stat.get("ci")
        if mean is None:
            add(f"| {label} | – | – |")
        else:
            ci_txt = f"{ci[0]:.3f} – {ci[1]:.3f}" if ci and ci[0] is not None else "–"
            add(f"| {label} | {mean:.3f} | {ci_txt} |")
    add(f"\n- Tính trên {summary['similarity_n']} dòng có claim ở CẢ HAI nhánh.")
    add("- Giống nhau CAO nghĩa là hai hệ hội tụ về cùng một câu; giống nhau THẤP nghĩa "
        "là chúng đi hai hướng khác hẳn. Bản thân con số không nói hệ nào đúng hơn.\n")

    add("## 3. Nhãn quan hệ\n")
    agree = summary["label_agreement"]
    add(f"- Cùng chấm được: **{agree['n']}** dòng")
    if agree["agreement"] is not None:
        add(f"- Đồng thuận thô: **{agree['agreement']:.1%}**")
    kappa = agree["cohen_kappa"]
    add(f"- Cohen's kappa: **{kappa:.3f}**" if kappa is not None
        else "- Cohen's kappa: **không xác định** (một nhánh chỉ ra một nhãn duy nhất)")
    add("\n| graph ↓ / rag → | supports | contradicts | irrelevant |")
    add("|---|---|---|---|")
    for ra in ("supports", "contradicts", "irrelevant"):
        row = agree["confusion"][ra]
        add(f"| **{ra}** | {row['supports']} | {row['contradicts']} | {row['irrelevant']} |")
    mc = summary["mcnemar"]
    add(f"\n- McNemar trên 'kết luận khác irrelevant': b={mc['b']}, c={mc['c']}, "
        f"p={mc['p']:.4f}")
    add(f"- Lực kiểm định phụ thuộc b+c={mc['b'] + mc['c']} (cần ≳25 mới đủ nhạy).\n")

    add("## 4. Đối chứng âm (trục Y — kỷ luật)\n")
    control = summary.get("negative_control")
    if not control:
        add("- **Không chạy.** Thiếu mục này thì các con số trên không phòng thủ được: "
            "không phân biệt được 'tìm ra bằng chứng thật' với 'dễ dãi hơn'.\n")
    else:
        add("Ghép mỗi câu bằng chứng với một claim NGẪU NHIÊN của cùng doanh nghiệp. "
            "Cặp ngẫu nhiên đúng ra phải bị chấm `irrelevant` — tỷ lệ làm đúng điều đó "
            "là specificity.\n")
        add("| nhánh | specificity | KTC 95% (Wilson) |")
        add("|---|---|---|")
        for arm in ("graph", "rag"):
            entry = control.get(arm) or {}
            k, n = entry.get("irrelevant", 0), entry.get("n", 0)
            ci = wilson_ci(k, n)
            ci_txt = f"{ci[0]:.1%} – {ci[1]:.1%}" if ci else "–"
            add(f"| {arm} | {k}/{n} = {k / n:.1%} |" f" {ci_txt} |" if n
                else f"| {arm} | – | – |")
        add("\n> Specificity thấp = nhánh đó gán quan hệ cho cả những cặp vô can. "
            "Coverage cao đi kèm specificity thấp là bẫy dễ dãi, phải báo cáo như "
            "đánh đổi kèm số, không được gọi là cải thiện.\n")

    add("## Giới hạn\n")
    add("- Không đo được accuracy: không có oracle. Mọi con số ở đây là tương đối "
        "giữa hai hệ.")
    add("- Nhánh graph khoanh claim theo issuer; nhánh rag lọc theo ticker. Cùng phạm vi.")
    add("- Cặp ngẫu nhiên trong mục 4 có thể tình cờ thật sự liên quan, nên specificity "
        "bị ước lượng THẤP hơn thực tế — lệch về phía bất lợi cho hệ, tức là an toàn.")
    return "\n".join(L)


# --------------------------------------------------------------------------
def main() -> int:
    ensure_utf8_stdout()
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--sheet-a", type=Path, default=SHEET_A)
    parser.add_argument("--graph", type=Path, default=RESOLVED_GRAPH)
    parser.add_argument("--out-dir", type=Path, default=SCRIPT_DIR)
    parser.add_argument("--out-name", default="graphrag_vs_rag")
    parser.add_argument("--limit", type=int, help="chỉ xử lý N dòng đầu")
    parser.add_argument("--top-k", type=int, default=20, help="ứng viên đưa cho reranker (nhánh rag)")
    parser.add_argument("--pool", type=int, default=100)
    parser.add_argument("--graph-top-k", type=int, default=10, help="ứng viên nhánh graph")
    parser.add_argument("--no-rerank", action="store_true", help="tắt LLM rerank ở nhánh rag")
    parser.add_argument("--no-negative-control", action="store_true",
                        help="bỏ đối chứng âm (KHÔNG khuyến khích — xem docstring)")
    parser.add_argument("--nc-sample", type=int, default=60,
                        help="đối chứng âm chạy trên N dòng lấy mẫu (0 = tất cả). "
                             "60 dòng đủ cho một ước lượng specificity kèm KTC Wilson, "
                             "mà rẻ hơn nhiều so với chạy cả 220")
    parser.add_argument("--seed", type=int, default=42, help="seed cho đối chứng âm")
    # Nút thắt là độ trễ mạng, không phải CPU: chạy tuần tự mất ~5,5 tiếng cho 220 dòng.
    # Kết quả KHÔNG phụ thuộc số luồng (gộp theo thứ tự dòng gốc, RNG gieo theo dòng).
    parser.add_argument("--workers", type=int, default=8, help="số dòng chạy song song")
    parser.add_argument("--model", help="model chấm nhãn (mặc định: theo endpoint trong .env)")
    parser.add_argument("--provider-order", default="openai")
    # 10 = mặc định của step07, nghĩa là 10 lệnh/PHÚT. Với 2 lệnh chấm nhãn mỗi dòng thì
    # riêng cái van đó đã ép cả lần chạy xuống ~5 dòng/phút. Endpoint đang dùng chịu được
    # hơn nhiều, nên nới ra — hạ lại nếu bị 429.
    parser.add_argument("--rate-limit", type=int, default=120)
    parser.add_argument("--openai-base-url",
                        help="mặc định lấy OPENAI_BASE_URL trong .env — cùng endpoint ragtest dùng")
    args = parser.parse_args()

    from esg_kg.core.llm import OPENAI_DEFAULT_MODEL
    from esg_kg.crosscheck.claims_vs_conduct import Adjudicator, node_text, props
    from ragtest.build_index import load_index
    from ragtest.company import load_aliases
    from ragtest.config import (CHAT_MODEL, DEFAULT_FINAL_K, EMBED_MODEL, INDEX_DIR,
                                ISSUER_REGISTRY, NEWS_PREPROCESSED)
    from ragtest.corpus import TICKERS
    from ragtest.llm import build_clients
    from ragtest.query import run_one
    from ragtest.rerank import LLMReranker
    from ragtest.retriever import HybridRetriever

    # Bộ chấm nhãn phải trỏ vào ĐÚNG endpoint đang có key. `.env` của repo này đặt
    # OPENAI_BASE_URL vào host GLM (ragtest cũng dùng nó), nên nếu để step07 mặc định
    # gọi OpenAI thật thì key GLM bị từ chối và toàn bộ nhãn rỗng — đúng lỗi lần chạy
    # thử đầu tiên. Lấy base_url từ .env, và khi có base_url thì model mặc định cũng
    # phải là model host đó phục vụ, không phải gpt-4o-mini.
    import os

    from ragtest.config import load_dotenv
    load_dotenv()
    base_url = args.openai_base_url or os.getenv("OPENAI_BASE_URL") or None
    model = args.model or (CHAT_MODEL if base_url else OPENAI_DEFAULT_MODEL)

    print("=" * 72)
    print(" Graph-RAG  vs  RAG thường")
    print("=" * 72)

    items = json.loads(Path(args.sheet_a).read_text(encoding="utf-8"))["items"]
    if args.limit:
        items = items[:args.limit]
    graph_data = json.loads(Path(args.graph).read_text(encoding="utf-8"))
    nodes = graph_data["nodes"]
    text_index = build_text_index(nodes)
    arm_graph = GraphArm(graph_data)
    print(f"\n[1/5] sheet_A          {len(items)} câu bằng chứng")
    print(f"[2/5] nhánh A (graph)  {len(nodes):,} node · {len(arm_graph.tickers())} issuer")

    corpus, embeddings, _meta = load_index(INDEX_DIR)
    aliases = load_aliases(ISSUER_REGISTRY, NEWS_PREPROCESSED)
    embedder = reranker = None
    try:
        clients = build_clients(CHAT_MODEL, EMBED_MODEL,
                                cache_path=INDEX_DIR / "embed_cache.json")
        embedder = clients["embed"]
        if not args.no_rerank:
            reranker = LLMReranker(clients["chat"], CHAT_MODEL)
    except Exception as exc:  # noqa: BLE001
        print(f"      CẢNH BÁO: không dựng được client ({exc}) — lùi về BM25 thuần")
    retriever = HybridRetriever(corpus, embeddings, embedder=embedder)
    print(f"[3/5] nhánh B (rag)    {len(corpus):,} câu claim · "
          f"{'bm25+dense' if embedder else 'bm25'}{' + rerank' if reranker else ''}")

    adjud = Adjudicator(model, args.rate_limit, [p.strip() for p in args.provider_order.split(",")],
                        base_url=base_url)
    if not adjud.enabled:
        print("LỖI: không có provider LLM nào khả dụng — cần OPENAI_API_KEY trong .env.")
        return 1
    print(f"[4/5] chấm nhãn        {model} @ {base_url or 'api.openai.com'} "
          f"(dùng CHUNG cho cả hai nhánh)")

    results_path = Path(args.out_dir) / f"{args.out_name}_rag_queries.jsonl"
    results_path.parent.mkdir(parents=True, exist_ok=True)
    if results_path.exists():
        results_path.unlink()
    query_args = argparse.Namespace(
        ticker=None, top_k=args.top_k, pool=args.pool, final_k=DEFAULT_FINAL_K,
        results=results_path, chat_model=CHAT_MODEL, embed_model=EMBED_MODEL,
        with_verdict=False)

    print(f"[5/5] đang chạy {len(items)} dòng · {args.workers} luồng…\n")
    stats = Counter()
    control = {"graph": Counter(), "rag": Counter()}

    def process(position: int, item: Dict[str, Any]) -> Dict[str, Any]:
        """
        Một dòng, độc lập hoàn toàn với các dòng khác.

        Chạy tuần tự thì hết ~5,5 tiếng cho 220 dòng — nút thắt là ĐỘ TRỄ MẠNG chứ không
        phải CPU, nên luồng ăn thua. An toàn vì: `RateLimiter` của Adjudicator có khoá
        riêng, `ChatClient` không giữ trạng thái giữa các lệnh, và mỗi dòng chỉ đọc chứ
        không ghi vào cấu trúc dùng chung (gộp số liệu làm sau, ở luồng chính).

        RNG của đối chứng âm gieo THEO DÒNG (`seed + position`) chứ không dùng chung một
        đối tượng: một RNG dùng chung sẽ cho kết quả phụ thuộc thứ tự luồng hoàn thành,
        tức là chạy lại ra số khác — mất tính tái lập, mà repo yêu cầu tái lập được.
        """
        evidence_text = item.get("evidence_text", "") or ""
        kind = item.get("evidence_kind_label")
        ticker = (item.get("claim_company") or "").upper()
        xi = match_node_index(evidence_text, kind, text_index, nodes)
        fields = row_fields(evidence_text, None if xi is None else nodes[xi])
        query = build_query(fields)
        meta = f"{kind or 'evidence'}, {fields.get('valid_from') or 'n/a'}"

        # ---- nhánh A: Graph-RAG ----
        # A  = top-1 theo thứ hạng của đồ thị (chưa rerank)
        # A' = top-1 sau khi CÙNG reranker của nhánh B sắp lại
        # Chỉ A' mới đem so với B: nếu một bên có reranker còn bên kia không, chênh lệch
        # đo được là chênh lệch của reranker chứ không phải của cơ chế truy xuất.
        graph_claim_norerank, graph_claim, graph_tier = "", "", ""
        if xi is not None:
            hits = arm_graph.retrieve(xi, ticker, top_k=args.graph_top_k)
            if hits:
                graph_claim_norerank = hits[0]["claim_text"]
                graph_tier = hits[0]["tier"]
                graph_claim = graph_claim_norerank
                if reranker is not None and len(hits) > 1:
                    # Reranker của ragtest đọc khoá `text`; nạp thêm meta để prompt của nó
                    # nhìn thấy cùng loại thông tin như ở nhánh B.
                    payload = [{"text": h["claim_text"], "ticker": ticker,
                                "source_pdf": "graph", "page": h["node_index"]} for h in hits]
                    with contextlib.redirect_stdout(io.StringIO()):
                        ordered, _ok = reranker.rerank(query, payload)
                    if ordered:
                        graph_claim = ordered[0].get("text", "") or graph_claim

        # ---- nhánh B: RAG thường ----
        # Namespace RIÊNG cho mỗi dòng: `run_one` đọc `args.ticker`, nên dùng chung một
        # đối tượng giữa các luồng sẽ khiến dòng này lọc theo mã của dòng khác.
        row_args = argparse.Namespace(**vars(query_args))
        row_args.ticker = ticker if ticker in TICKERS else None
        with contextlib.redirect_stdout(io.StringIO()):
            record = run_one(query, retriever, aliases, reranker, None, row_args)
        candidates = record.get("candidates") or []
        rag_claim = candidates[0].get("text", "") if candidates else ""

        # ---- nhãn: cùng một Adjudicator ----
        graph_relation = adjudicate_label(adjud, graph_claim, evidence_text, meta)
        rag_relation = adjudicate_label(adjud, rag_claim, evidence_text, meta)

        sim = text_similarity(graph_claim, rag_claim)
        cos = None
        if embedder is not None and graph_claim and rag_claim:
            try:
                vectors = embedder.embed([graph_claim, rag_claim])
                cos = cosine(list(vectors[0]), list(vectors[1]))
            except Exception:  # noqa: BLE001 — nhúng hỏng không được làm hỏng cả dòng
                cos = None

        row = {
            "_position": position,
            "pair_id": item.get("pair_id", ""), "ticker": ticker,
            "evidence_text": evidence_text,
            "graph_claim": graph_claim, "graph_tier": graph_tier,
            "graph_claim_norerank": graph_claim_norerank,
            "graph_rerank_changed": (graph_claim != graph_claim_norerank) if graph_claim_norerank else "",
            "graph_relation": graph_relation,
            "rag_claim": rag_claim, "rag_relation": rag_relation,
            **{k: round(v, 4) for k, v in sim.items()},
            "cosine": "" if cos is None else round(cos, 4),
            "labels_agree": (graph_relation == rag_relation) if (graph_relation and rag_relation) else "",
        }

        # ---- đối chứng âm: claim ngẫu nhiên cùng doanh nghiệp ----
        # Lấy mẫu N dòng đầu thay vì cả 220: specificity là một tỷ lệ nhị phân, 60 mẫu
        # đủ cho một khoảng Wilson dùng được, mà tiết kiệm 2 lệnh LLM cho mỗi dòng còn lại.
        row["_control"] = []
        run_control = (not args.no_negative_control
                       and (args.nc_sample <= 0 or position <= args.nc_sample))
        if run_control:
            pool = arm_graph.claims_for_ticker(ticker)
            if pool:
                decoy = node_text(nodes[random.Random(args.seed + position).choice(pool)])
                for arm_name, real_claim in (("graph", graph_claim), ("rag", rag_claim)):
                    if not real_claim:
                        continue
                    verdict = adjudicate_label(adjud, decoy, evidence_text, meta)
                    if verdict:
                        row["_control"].append((arm_name, verdict == "irrelevant"))
        return row

    # Gộp ở luồng chính, theo thứ tự dòng gốc — nên kết quả không phụ thuộc luồng nào
    # xong trước.
    done = 0
    with ThreadPoolExecutor(max_workers=args.workers) as pool_exec:
        futures = {pool_exec.submit(process, i, it): i
                   for i, it in enumerate(items, start=1)}
        collected: Dict[int, Dict[str, Any]] = {}
        for future in as_completed(futures):
            position = futures[future]
            try:
                collected[position] = future.result()
            except Exception as exc:  # noqa: BLE001 — một dòng hỏng không giết cả lần chạy
                print(f"      LỖI dòng {position}: {type(exc).__name__}: {exc}")
            done += 1
            if done % 10 == 0 or done == len(items):
                print(f"      {done}/{len(items)}")

    rows: List[Dict[str, Any]] = [collected[i] for i in sorted(collected)]
    for row in rows:
        stats["graph_answered"] += bool(row["graph_claim"])
        stats["rag_answered"] += bool(row["rag_claim"])
        if row["graph_tier"]:
            stats["tier_" + row["graph_tier"]] += 1
        if row["graph_claim_norerank"]:
            stats["graph_rerank_changed"] += (row["graph_claim"] != row["graph_claim_norerank"])
        for arm_name, is_irrelevant in row.pop("_control", []):
            control[arm_name]["n"] += 1
            control[arm_name]["irrelevant"] += is_irrelevant
        row.pop("_position", None)

    if embedder is not None:
        with contextlib.suppress(Exception):
            embedder.save_cache()

    # ---- tổng hợp ----
    both = [r for r in rows if r["graph_claim"] and r["rag_claim"]]
    similarity: Dict[str, Any] = {}
    for key in ("token_f1", "token_precision", "token_recall", "jaccard", "rouge_l"):
        values = [r[key] for r in both]
        similarity[key] = {"mean": (sum(values) / len(values)) if values else None,
                           "ci": list(bootstrap_ci(values, reps=2000, seed=args.seed))}
    cosines = [r["cosine"] for r in both if r["cosine"] != ""]
    similarity["cosine"] = {"mean": (sum(cosines) / len(cosines)) if cosines else None,
                            "ci": list(bootstrap_ci(cosines, reps=2000, seed=args.seed))}

    labelled = [r for r in rows if r["graph_relation"] and r["rag_relation"]]
    agreement = label_agreement([r["graph_relation"] for r in labelled],
                                [r["rag_relation"] for r in labelled])

    # McNemar trên "kết luận khác irrelevant" — đo nhánh nào dễ dãi hơn (trục Y của §2)
    b = sum(1 for r in labelled if r["graph_relation"] != "irrelevant" and r["rag_relation"] == "irrelevant")
    c = sum(1 for r in labelled if r["graph_relation"] == "irrelevant" and r["rag_relation"] != "irrelevant")

    summary = {
        "n_items": len(rows),
        "model": model,
        "graph_answered": stats["graph_answered"],
        "rag_answered": stats["rag_answered"],
        "graph_tiers": {"indicator": stats["tier_indicator"],
                        "token_overlap": stats["tier_token_overlap"]},
        "reranked": reranker is not None,
        "graph_rerank_changed": stats["graph_rerank_changed"],
        "nc_sample": args.nc_sample,
        "similarity_n": len(both),
        "similarity": similarity,
        "label_agreement": agreement,
        "mcnemar": {"b": b, "c": c, "p": mcnemar_exact(b, c)},
        "negative_control": None if args.no_negative_control else
            {arm: dict(counter) for arm, counter in control.items()},
        "seed": args.seed,
        "no_ground_truth": True,
    }

    out_dir = Path(args.out_dir)
    print("\nĐã ghi:")
    write_outputs(rows, summary, out_dir, args.out_name)
    report_path = out_dir / f"{args.out_name}_report.md"
    report_path.write_text(render_report(summary), encoding="utf-8")
    print(f"  {report_path}")

    print(f"\nTổng kết {len(rows)} dòng")
    print(f"  graph ra claim : {stats['graph_answered']}/{len(rows)}")
    print(f"  rag   ra claim : {stats['rag_answered']}/{len(rows)}")
    print(f"  cùng chấm nhãn : {agreement['n']}")
    if agreement["agreement"] is not None:
        print(f"  đồng thuận nhãn: {agreement['agreement']:.1%}"
              f"  kappa={agreement['cohen_kappa'] if agreement['cohen_kappa'] is None else round(agreement['cohen_kappa'], 3)}")
    if similarity["token_f1"]["mean"] is not None:
        print(f"  token F1 TB    : {similarity['token_f1']['mean']:.3f}")
    if similarity["cosine"]["mean"] is not None:
        print(f"  cosine TB      : {similarity['cosine']['mean']:.3f}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
