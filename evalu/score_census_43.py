#!/usr/bin/env python3
"""
score_census_43.py — the reproducible source of the 43-pair census numbers
quoted in capstone_report/main.tex (S:experiment-expert), reported for issue #17.

Why this script exists
-----------------------
build_census_43.py wrote the blind sheets (evalu/out/annotation/census_43/) and the two
experts filled them in Excel (sheetA_43pairs_filled.xlsx / sheetB_43pairs_filled.xlsx,
committed at the repo root). Every number in main.tex's S4.4 rewrite -- precision by
annotator and by kind, the confusion tables, Cohen's kappa, Gwet's AC1, the
company/claim breakdown -- was first computed with one-off interactive commands. Per
docs/ANNOTATION_RESULTS.md's own rule ("Dung tu tinh lai bang tay"), that is not
acceptable as the permanent record: this script is the thing that regenerates them,
so a reader can check the thesis against evalu/annotation.py's score() and evalu/iaa.py
directly rather than trusting a transcript.

What it does NOT do: touch Neo4j. The one-time check that these 43 pairs are exactly
what the live advisory layer serves (37 llm_supports + 6 llm_contradicts, matched pair
for pair by (claim_id, evidence_text)) was a separate, interactive verification against
a running Neo4j instance -- not reproducible offline and not this script's job. This
script only re-derives what's on disk: the filled sheets against the answer key baked
into evalu/out/annotation/census_43/sheet_A.json / sheet_B.json at build time.

Offline: reads only the filled xlsx sheets + the census JSON keys already on disk. No
LLM, no Neo4j, no network.

    python evalu/score_census_43.py
"""
from __future__ import annotations

import argparse
import json
import math
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Tuple

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT))

import openpyxl  # noqa: E402

from evalu.annotation import KIND_TO_RELATION, LABELS, score  # noqa: E402
from evalu.iaa import cohen_kappa, gwet_ac1, landis_koch_label  # noqa: E402

CENSUS_DIR = REPO_ROOT / "evalu" / "out" / "annotation" / "census_43"


def read_xlsx_annotations(path: Path) -> List[Dict[str, Any]]:
    """Read a filled sheet back from Excel, same field set as evalu.annotation's CSV."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(h).strip() for h in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:]]


def _wilson_ci(k: int, n: int, z: float = 1.96) -> Tuple[float, float]:
    if n == 0:
        return (0.0, 0.0)
    p = k / n
    denom = 1 + z * z / n
    centre = (p + z * z / (2 * n)) / denom
    half = z * math.sqrt(p * (1 - p) / n + z * z / (4 * n * n)) / denom
    return (max(0.0, centre - half), min(1.0, centre + half))


def _pct(p: float) -> str:
    return f"{p * 100:.1f}%"


def confusion_table(sheet: Dict[str, Any], annotations: List[Dict[str, Any]]) -> Dict[str, Counter]:
    """System-asserted relation (row) -> human label (column), matching tab:expert-confusion."""
    key = sheet.get("_key", {})
    out: Dict[str, Counter] = defaultdict(Counter)
    for a in annotations:
        pid = (a.get("pair_id") or "").strip()
        rel = (a.get("relation") or "").strip().lower()
        meta = key.get(pid)
        if not pid or rel not in LABELS or meta is None:
            continue
        claimed = KIND_TO_RELATION.get(meta.get("system_kind"))
        if claimed is None:
            continue
        out[claimed][rel] += 1
    return out


def print_precision_report(name: str, sheet: Dict[str, Any], annotations: List[Dict[str, Any]]) -> None:
    result = score(sheet, annotations)
    prec = result["precision"]
    print(f"=== {name} ===")
    o = prec["overall"]
    lo, hi = _wilson_ci(o["correct"], o["n"])
    print(f"  overall       {o['correct']}/{o['n']} = {_pct(o['value'])}  Wilson95%[{_pct(lo)}, {_pct(hi)}]")
    for kind in ("supporting_evidence", "contradicting_evidence"):
        k = prec["by_kind"].get(kind)
        if not k:
            continue
        lo, hi = _wilson_ci(k["correct"], k["n"])
        print(f"  {kind:24s} {k['correct']}/{k['n']} = {_pct(k['value'])}  Wilson95%[{_pct(lo)}, {_pct(hi)}]")
    sco = prec["same_company_only"]
    if sco["n"]:
        print(f"  same_company_only  {sco['correct']}/{sco['n']} = {_pct(sco['value'])}  "
              f"(n={sco['n']} -- only pairs the annotator marked about_claim_company=Yes)")
    confusion = confusion_table(sheet, annotations)
    print("  confusion (system asserts -> human label):")
    for claimed in ("supports", "contradicts"):
        row = confusion.get(claimed, Counter())
        n = sum(row.values())
        print(f"    {claimed:12s} (n={n})  " +
              "  ".join(f"{lbl}={row.get(lbl, 0)}" for lbl in LABELS))
    print()


def about_claim_company_completeness(annotations: List[Dict[str, Any]]) -> Tuple[int, int]:
    """(filled, total) — how many of the 43 rows carry an about_claim_company answer."""
    filled = sum(1 for a in annotations
                 if str(a.get("about_claim_company") or "").strip().lower() not in ("", "none"))
    return filled, len(annotations)


def population_breakdown(sheet: Dict[str, Any]) -> None:
    key = sheet.get("_key", {})
    by_company: Counter = Counter()
    claim_ids = set()
    for pid, meta in key.items():
        by_company[meta.get("claim_company")] += 1
    # claim_ids not recoverable from _key alone (it doesn't carry claim_id) --
    # reported separately by build_census_43.py's own stdout at build time.
    print("Population by company (from the answer key):", dict(by_company))


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--sheet-a-xlsx", type=Path, default=REPO_ROOT / "sheetA_43pairs_filled.xlsx")
    ap.add_argument("--sheet-b-xlsx", type=Path, default=REPO_ROOT / "sheetB_43pairs_filled.xlsx")
    args = ap.parse_args()

    sheet_a = json.loads((CENSUS_DIR / "sheet_A.json").read_text(encoding="utf-8"))
    sheet_b = json.loads((CENSUS_DIR / "sheet_B.json").read_text(encoding="utf-8"))

    ann_a = read_xlsx_annotations(args.sheet_a_xlsx)
    ann_b = read_xlsx_annotations(args.sheet_b_xlsx)

    population_breakdown(sheet_a)
    print()
    print_precision_report("Annotator A (Thai Anh Tuan)", sheet_a, ann_a)
    print_precision_report("Annotator B (Do Kim Ngoc)", sheet_b, ann_b)

    filled_a, total_a = about_claim_company_completeness(ann_a)
    filled_b, total_b = about_claim_company_completeness(ann_b)
    print(f"about_claim_company filled: A {filled_a}/{total_a}, B {filled_b}/{total_b} "
          "(expect 19/43 -- only carried-over pairs from the earlier 200-pair round; "
          "the 24 pairs new to this census were left blank by both annotators)")
    print()

    a_by_pid = {r["pair_id"]: (r.get("relation") or "").strip().lower() for r in ann_a}
    b_by_pid = {r["pair_id"]: (r.get("relation") or "").strip().lower() for r in ann_b}
    shared = sorted(set(a_by_pid) & set(b_by_pid))
    matrix = [[a_by_pid[p], b_by_pid[p]] for p in shared]
    agree = sum(1 for x, y in matrix if x == y)
    kappa = cohen_kappa(matrix)
    ac1 = gwet_ac1(matrix)
    print(f"Inter-annotator agreement, n={len(matrix)}:")
    print(f"  raw agreement  {agree}/{len(matrix)} = {_pct(agree / len(matrix))}")
    print(f"  Cohen's kappa  {kappa:.3f} ({landis_koch_label(kappa)})")
    print(f"  Gwet's AC1     {ac1:.3f} ({landis_koch_label(ac1)})")
    disagreements = [p for p in shared if a_by_pid[p] != b_by_pid[p]]
    for p in disagreements:
        print(f"  disagreement: pair_id={p}  A={a_by_pid[p]}  B={b_by_pid[p]}")


if __name__ == "__main__":
    main()
